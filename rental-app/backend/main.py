"""建設レンタル・廃棄物処理管理アプリ API"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import secrets
from datetime import date, datetime

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles

from . import db as D
from .auth import admin_user, create_session, current_user, hash_password, verify_password
from .db import audit, get_db, get_pref, now_str, set_pref
from .pricing import calc_rental_charge

app = FastAPI(title="建設レンタル・廃棄物処理管理")

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC = os.path.join(BASE, "static")
UPLOADS = os.path.join(STATIC, "uploads")
os.makedirs(UPLOADS, exist_ok=True)

D.init_db()

from .seed_import import seed as _seed_import  # noqa: E402
_seed_import(UPLOADS)


def err(errors: dict, status: int = 422):
    return JSONResponse({"errors": errors}, status_code=status)


def parse_date(s, field, errors, label):
    if not s:
        errors[field] = f"{label}を選んでください"
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        errors[field] = f"{label}の形式が正しくありません"
        return None


def device_of(request: Request) -> str:
    ua = request.headers.get("user-agent", "")
    if "iPad" in ua:
        return "iPad"
    if "iPhone" in ua:
        return "iPhone"
    return "PC/その他"


# ---------------------------------------------------------------- 認証・初期設定
@app.get("/api/state")
def state(request: Request):
    conn = get_db()
    setup_done = conn.execute("SELECT COUNT(*) c FROM users WHERE role='admin'").fetchone()["c"] > 0
    user = None
    token = request.cookies.get("session")
    if token:
        row = conn.execute(
            "SELECT u.* FROM sessions s JOIN users u ON u.id=s.user_id "
            "WHERE s.token=? AND s.expires_at>=?", (token, now_str())).fetchone()
        if row and row["active"]:
            user = {k: row[k] for k in ("id", "role", "name", "display_name",
                                        "login_id", "must_change_password")}
    conn.close()
    return {"setup_done": setup_done, "user": user}


@app.post("/api/setup")
async def setup(request: Request):
    body = await request.json()
    conn = get_db()
    try:
        if conn.execute("SELECT COUNT(*) c FROM users WHERE role='admin'").fetchone()["c"]:
            raise HTTPException(400, "初期設定はすでに完了しています")
        errors = {}
        for f, label in [("admin_name", "管理者名"), ("company_name", "会社名"),
                         ("email", "メールアドレス"), ("password", "パスワード")]:
            if not (body.get(f) or "").strip():
                errors[f] = f"{label}を入力してください"
        if body.get("password") and len(body["password"]) < 8:
            errors["password"] = "パスワードは8文字以上にしてください"
        if errors:
            return err(errors)
        login_id = (body.get("login_id") or body["email"]).strip()
        conn.execute(
            "INSERT INTO users(role,name,display_name,login_id,email,password_hash,created_at)"
            " VALUES('admin',?,?,?,?,?,?)",
            (body["admin_name"], body["admin_name"], login_id, body["email"],
             hash_password(body["password"]), now_str()))
        conn.execute("INSERT INTO company(id,company_name,admin_name,created_at) VALUES(1,?,?,?)",
                     (body["company_name"], body["admin_name"], now_str()))
        uid = conn.execute("SELECT id FROM users WHERE login_id=?", (login_id,)).fetchone()["id"]
        audit(conn, {"id": uid, "name": body["admin_name"]}, "初期設定", "company", 1,
              after={"company_name": body["company_name"]}, device=device_of(request))
        token = create_session(conn, uid)
        conn.commit()
        resp = JSONResponse({"ok": True})
        resp.set_cookie("session", token, httponly=True, samesite="lax", max_age=86400 * 30)
        return resp
    finally:
        conn.close()


@app.post("/api/login")
async def login(request: Request):
    body = await request.json()
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM users WHERE login_id=?",
                           ((body.get("login_id") or "").strip(),)).fetchone()
        if not row or not verify_password(body.get("password") or "", row["password_hash"]):
            return err({"login": "ログインIDまたはパスワードが違います"}, 401)
        if not row["active"]:
            return err({"login": "このアカウントは停止されています。管理者に連絡してください"}, 403)
        token = create_session(conn, row["id"])
        audit(conn, dict(row), "ログイン", "user", row["id"], device=device_of(request))
        conn.commit()
        resp = JSONResponse({"ok": True, "user": {
            "id": row["id"], "role": row["role"], "name": row["name"],
            "display_name": row["display_name"],
            "must_change_password": row["must_change_password"]}})
        resp.set_cookie("session", token, httponly=True, samesite="lax", max_age=86400 * 30)
        return resp
    finally:
        conn.close()


@app.post("/api/logout")
def logout(request: Request):
    token = request.cookies.get("session")
    if token:
        conn = get_db()
        conn.execute("DELETE FROM sessions WHERE token=?", (token,))
        conn.commit()
        conn.close()
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("session")
    return resp


@app.post("/api/change_password")
async def change_password(request: Request, user=Depends(current_user)):
    body = await request.json()
    new = body.get("new_password") or ""
    if len(new) < 8:
        return err({"new_password": "新しいパスワードは8文字以上にしてください"})
    conn = get_db()
    try:
        row = conn.execute("SELECT password_hash FROM users WHERE id=?", (user["id"],)).fetchone()
        if not verify_password(body.get("old_password") or "", row["password_hash"]):
            return err({"old_password": "現在のパスワードが違います"})
        conn.execute("UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?",
                     (hash_password(new), user["id"]))
        audit(conn, user, "パスワード変更", "user", user["id"], device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


# ---------------------------------------------------------------- 従業員管理（管理者）
@app.get("/api/users")
def list_users(user=Depends(admin_user)):
    conn = get_db()
    rows = [dict(r) for r in conn.execute(
        "SELECT id,role,name,display_name,login_id,email,active,must_change_password,"
        "created_at FROM users ORDER BY id")]
    conn.close()
    return rows


@app.post("/api/users")
async def create_user(request: Request, user=Depends(admin_user)):
    body = await request.json()
    errors = {}
    for f, label in [("name", "氏名"), ("display_name", "表示名"),
                     ("login_id", "ログインID"), ("temp_password", "仮パスワード")]:
        if not (body.get(f) or "").strip():
            errors[f] = f"{label}を入力してください"
    if errors:
        return err(errors)
    conn = get_db()
    try:
        if conn.execute("SELECT 1 FROM users WHERE login_id=?", (body["login_id"],)).fetchone():
            return err({"login_id": "このログインIDはすでに使われています"})
        conn.execute(
            "INSERT INTO users(role,name,display_name,login_id,email,password_hash,"
            "must_change_password,created_at) VALUES('employee',?,?,?,?,?,1,?)",
            (body["name"], body["display_name"], body["login_id"].strip(),
             body.get("email", ""), hash_password(body["temp_password"]), now_str()))
        uid = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        if body.get("sites"):
            set_pref(conn, uid, "assigned_sites", body["sites"])
        audit(conn, user, "従業員追加", "user", uid,
              after={"name": body["name"], "login_id": body["login_id"]},
              device=device_of(request))
        conn.commit()
        return {"ok": True, "id": uid}
    finally:
        conn.close()


@app.post("/api/users/{uid}/toggle_active")
def toggle_active(uid: int, request: Request, user=Depends(admin_user)):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        if not row:
            raise HTTPException(404, "利用者が見つかりません")
        if row["role"] == "admin":
            raise HTTPException(400, "管理者アカウントは停止できません")
        conn.execute("UPDATE users SET active=? WHERE id=?", (0 if row["active"] else 1, uid))
        audit(conn, user, "利用停止" if row["active"] else "利用再開", "user", uid,
              device=device_of(request))
        conn.commit()
        return {"ok": True, "active": 0 if row["active"] else 1}
    finally:
        conn.close()


@app.post("/api/users/{uid}/reset_password")
async def reset_password(uid: int, request: Request, user=Depends(admin_user)):
    body = await request.json()
    if not (body.get("temp_password") or "").strip():
        return err({"temp_password": "仮パスワードを入力してください"})
    conn = get_db()
    try:
        conn.execute("UPDATE users SET password_hash=?, must_change_password=1 WHERE id=?",
                     (hash_password(body["temp_password"]), uid))
        audit(conn, user, "仮パスワード再発行", "user", uid, device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


# ---------------------------------------------------------------- メタ情報
@app.get("/api/meta")
def meta(user=Depends(current_user)):
    return {
        "waste_types": D.WASTE_TYPES, "waste_types_top": D.WASTE_TYPES_TOP,
        "waste_units": D.WASTE_UNITS, "extension_reasons": D.EXTENSION_REASONS,
        "photo_categories": D.PHOTO_CATEGORIES,
    }


# ---------------------------------------------------------------- 現場
@app.get("/api/sites")
def list_sites(q: str = "", contractor: str = "", user=Depends(current_user)):
    conn = get_db()
    sql = ("SELECT s.*, u.display_name creator FROM sites s "
           "LEFT JOIN users u ON u.id=s.created_by WHERE s.deleted=0")
    args = []
    if q:
        sql += " AND s.name LIKE ?"
        args.append(f"%{q}%")
    if contractor:
        sql += " AND s.contractor LIKE ?"
        args.append(f"%{contractor}%")
    rows = [dict(r) for r in conn.execute(sql + " ORDER BY s.id DESC", args)]
    conn.close()
    return rows


@app.get("/api/sites/pickdata")
def site_pickdata(user=Depends(current_user)):
    """現場選択画面用: 最近使用・前回・お気に入り・施工中。"""
    conn = get_db()
    try:
        recent_ids = get_pref(conn, user["id"], "recent_sites", [])
        last_site = get_pref(conn, user["id"], "last_site")
        favorites = get_pref(conn, user["id"], "favorite_sites", [])
        sites = {r["id"]: dict(r) for r in conn.execute(
            "SELECT * FROM sites WHERE deleted=0")}
        return {
            "last_site": sites.get(last_site),
            "recent": [sites[i] for i in recent_ids if i in sites],
            "favorites": [sites[i] for i in favorites if i in sites],
            "active": [s for s in sites.values() if s["status"] == "active"],
        }
    finally:
        conn.close()


@app.post("/api/sites")
async def create_site(request: Request, user=Depends(current_user)):
    """現場の登録（従業員も可）。同名の現場があれば重複登録せずそれを返す。"""
    body = await request.json()
    name = (body.get("name") or "").strip()
    if not name:
        return err({"name": "現場名を入力してください"})
    conn = get_db()
    try:
        dup = conn.execute("SELECT * FROM sites WHERE name=? AND deleted=0", (name,)).fetchone()
        if dup:
            return {"ok": True, "id": dup["id"], "existing": True,
                    "site": dict(dup)}
        conn.execute("INSERT INTO sites(name,contractor,status,address,created_by,created_at)"
                     " VALUES(?,?,?,?,?,?)",
                     (name, body.get("contractor", ""), body.get("status", "active"),
                      body.get("address", ""), user["id"], now_str()))
        sid = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        audit(conn, user, "現場登録", "site", sid, after=body, device=device_of(request))
        conn.commit()
        site = dict(conn.execute("SELECT * FROM sites WHERE id=?", (sid,)).fetchone())
        return {"ok": True, "id": sid, "existing": False, "site": site}
    finally:
        conn.close()


@app.put("/api/sites/{sid}")
async def update_site(sid: int, request: Request, user=Depends(admin_user)):
    body = await request.json()
    conn = get_db()
    try:
        before = conn.execute("SELECT * FROM sites WHERE id=?", (sid,)).fetchone()
        if not before:
            raise HTTPException(404, "現場が見つかりません")
        conn.execute("UPDATE sites SET name=?,contractor=?,status=?,address=? WHERE id=?",
                     (body.get("name", before["name"]), body.get("contractor", before["contractor"]),
                      body.get("status", before["status"]), body.get("address", before["address"]), sid))
        audit(conn, user, "現場編集", "site", sid, before=dict(before), after=body,
              device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.delete("/api/sites/{sid}")
def delete_site(sid: int, request: Request, user=Depends(admin_user)):
    conn = get_db()
    try:
        before = conn.execute("SELECT * FROM sites WHERE id=?", (sid,)).fetchone()
        conn.execute("UPDATE sites SET deleted=1 WHERE id=?", (sid,))
        audit(conn, user, "現場削除", "site", sid, before=dict(before) if before else None,
              device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.post("/api/sites/{sid}/favorite")
def toggle_fav_site(sid: int, user=Depends(current_user)):
    conn = get_db()
    try:
        favs = get_pref(conn, user["id"], "favorite_sites", [])
        favs = [f for f in favs if f != sid] if sid in favs else favs + [sid]
        set_pref(conn, user["id"], "favorite_sites", favs)
        conn.commit()
        return {"ok": True, "favorites": favs}
    finally:
        conn.close()


def touch_site_usage(conn, user_id: int, site_id: int):
    set_pref(conn, user_id, "last_site", site_id)
    recent = get_pref(conn, user_id, "recent_sites", [])
    recent = [site_id] + [s for s in recent if s != site_id]
    set_pref(conn, user_id, "recent_sites", recent[:10])


# ---------------------------------------------------------------- 業者
@app.get("/api/vendors")
def list_vendors(kind: str = "", user=Depends(current_user)):
    conn = get_db()
    sql = "SELECT * FROM vendors WHERE deleted=0"
    args = []
    if kind:
        sql += " AND kind=?"
        args.append(kind)
    rows = [dict(r) for r in conn.execute(sql + " ORDER BY name", args)]
    conn.close()
    return rows


@app.post("/api/vendors")
async def create_vendor(request: Request, user=Depends(current_user)):
    body = await request.json()
    kind = body.get("kind")
    if kind not in ("rental", "hauler", "disposal"):
        return err({"kind": "業者の種類を選んでください"})
    if not (body.get("name") or "").strip():
        return err({"name": "業者名を入力してください"})
    conn = get_db()
    try:
        row = conn.execute("SELECT id FROM vendors WHERE kind=? AND name=? AND deleted=0",
                           (kind, body["name"].strip())).fetchone()
        if row:
            return {"ok": True, "id": row["id"]}
        conn.execute("INSERT INTO vendors(kind,name) VALUES(?,?)", (kind, body["name"].strip()))
        vid = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        audit(conn, user, "業者登録", "vendor", vid, after=body, device=device_of(request))
        conn.commit()
        return {"ok": True, "id": vid}
    finally:
        conn.close()


@app.delete("/api/vendors/{vid}")
def delete_vendor(vid: int, request: Request, user=Depends(admin_user)):
    conn = get_db()
    conn.execute("UPDATE vendors SET deleted=1 WHERE id=?", (vid,))
    audit(conn, user, "業者削除", "vendor", vid, device=device_of(request))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---------------------------------------------------------------- 料金マスター
@app.get("/api/price_master")
def list_price_master(q: str = "", user=Depends(current_user)):
    conn = get_db()
    sql = "SELECT * FROM price_master WHERE active=1"
    args = []
    if q:
        sql += " AND (name LIKE ? OR code LIKE ? OR spec LIKE ?)"
        args += [f"%{q}%"] * 3
    rows = [dict(r) for r in conn.execute(sql + " ORDER BY name", args)]
    conn.close()
    return rows


def _price_fields(body, errors):
    vals = {}
    for f, label in [("daily_rate", "日割単価"), ("monthly_rate", "月極単価"),
                     ("basic_fee", "基本料"), ("support_per_day", "サポート料／日"),
                     ("damage_per_day", "賠償対策費／日")]:
        v = body.get(f)
        if v in (None, ""):
            vals[f] = None
            continue
        try:
            vals[f] = int(v)
            if vals[f] < 0:
                errors[f] = f"{label}は0以上で入力してください"
        except (TypeError, ValueError):
            errors[f] = f"{label}は数字で入力してください"
    return vals


@app.post("/api/price_master")
async def create_price(request: Request, user=Depends(admin_user)):
    body = await request.json()
    errors = {}
    if not (body.get("name") or "").strip():
        errors["name"] = "品名を入力してください"
    vals = _price_fields(body, errors)
    if errors:
        return err(errors)
    needs_review = 1 if any(vals[f] is None for f in vals) else 0
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO price_master(code,name,spec,daily_rate,monthly_rate,basic_fee,"
            "support_per_day,damage_per_day,needs_review,source_image,created_at)"
            " VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            (body.get("code", ""), body["name"].strip(), body.get("spec", ""),
             vals["daily_rate"], vals["monthly_rate"], vals["basic_fee"],
             vals["support_per_day"], vals["damage_per_day"], needs_review,
             body.get("source_image", ""), now_str()))
        pid = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        audit(conn, user, "料金マスター登録", "price_master", pid, after=body,
              device=device_of(request))
        conn.commit()
        return {"ok": True, "id": pid, "needs_review": needs_review}
    finally:
        conn.close()


@app.put("/api/price_master/{pid}")
async def update_price(pid: int, request: Request, user=Depends(admin_user)):
    body = await request.json()
    errors = {}
    vals = _price_fields(body, errors)
    if errors:
        return err(errors)
    conn = get_db()
    try:
        before = conn.execute("SELECT * FROM price_master WHERE id=?", (pid,)).fetchone()
        if not before:
            raise HTTPException(404, "料金マスターが見つかりません")
        needs_review = 1 if any(vals[f] is None for f in vals) else 0
        conn.execute(
            "UPDATE price_master SET code=?,name=?,spec=?,daily_rate=?,monthly_rate=?,"
            "basic_fee=?,support_per_day=?,damage_per_day=?,needs_review=?,active=? WHERE id=?",
            (body.get("code", before["code"]), body.get("name", before["name"]),
             body.get("spec", before["spec"]), vals["daily_rate"], vals["monthly_rate"],
             vals["basic_fee"], vals["support_per_day"], vals["damage_per_day"],
             needs_review, body.get("active", before["active"]), pid))
        audit(conn, user, "料金マスター変更", "price_master", pid,
              before=dict(before), after=body, device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.get("/api/items/pickdata")
def item_pickdata(user=Depends(current_user)):
    conn = get_db()
    try:
        recent = get_pref(conn, user["id"], "recent_items", [])
        favs = get_pref(conn, user["id"], "favorite_items", [])
        items = {r["id"]: dict(r) for r in conn.execute(
            "SELECT * FROM price_master WHERE active=1")}
        return {"recent": [items[i] for i in recent if i in items],
                "favorites": [items[i] for i in favs if i in items]}
    finally:
        conn.close()


@app.post("/api/items/{iid}/favorite")
def toggle_fav_item(iid: int, user=Depends(current_user)):
    conn = get_db()
    try:
        favs = get_pref(conn, user["id"], "favorite_items", [])
        favs = [f for f in favs if f != iid] if iid in favs else favs + [iid]
        set_pref(conn, user["id"], "favorite_items", favs)
        conn.commit()
        return {"ok": True, "favorites": favs}
    finally:
        conn.close()


# ---------------------------------------------------------------- レンタル
def rental_snapshot_estimate(row: dict) -> dict | None:
    try:
        end = row.get("returned_date") or row["due_date"]
        c = calc_rental_charge(
            daily_rate=row["daily_rate"] or 0, monthly_rate=row["monthly_rate"] or 0,
            basic_fee=row["basic_fee"] or 0, support_per_day=row["support_per_day"] or 0,
            damage_per_day=row["damage_per_day"] or 0, qty=row["qty"],
            start=date.fromisoformat(row["start_date"]), end=date.fromisoformat(end))
        return c.to_dict()
    except (ValueError, KeyError):
        return None


@app.post("/api/rentals")
async def create_rental(request: Request, user=Depends(current_user)):
    body = await request.json()
    errors = {}
    site_id = body.get("site_id")
    if not site_id:
        errors["site_id"] = "現場を選んでください"
    if not body.get("vendor_id") and not (body.get("vendor_name") or "").strip():
        errors["vendor_id"] = "レンタル業者を選んでください"
    item_id = body.get("item_id")
    item_name = (body.get("item_name") or "").strip()
    if not item_id and not item_name:
        errors["item_id"] = "商品を選んでください"
    try:
        qty = int(body.get("qty") or 0)
        if qty < 1:
            errors["qty"] = "数量は1以上で入力してください"
    except (TypeError, ValueError):
        qty = 0
        errors["qty"] = "数量は数字で入力してください"
    start = parse_date(body.get("start_date"), "start_date", errors, "レンタル開始日")
    due = parse_date(body.get("due_date"), "due_date", errors, "返却予定日")
    if start and due and due < start:
        errors["due_date"] = "返却予定日は開始日以降にしてください"
    if errors:
        return err(errors)

    conn = get_db()
    try:
        # 二重登録防止（クライアント生成キー）
        ck = body.get("client_key")
        if ck:
            dup = conn.execute("SELECT id FROM rentals WHERE client_key=?", (ck,)).fetchone()
            if dup:
                return {"ok": True, "id": dup["id"], "duplicate": True}
        vendor_id = body.get("vendor_id")
        if not vendor_id and body.get("vendor_name"):
            r = conn.execute("SELECT id FROM vendors WHERE kind='rental' AND name=? AND deleted=0",
                             (body["vendor_name"].strip(),)).fetchone()
            if r:
                vendor_id = r["id"]
            else:
                conn.execute("INSERT INTO vendors(kind,name) VALUES('rental',?)",
                             (body["vendor_name"].strip(),))
                vendor_id = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        # 料金は料金マスターから自動取得（従業員は編集不可）
        snap = {"daily_rate": 0, "monthly_rate": 0, "basic_fee": 0,
                "support_per_day": 0, "damage_per_day": 0, "spec": body.get("spec", ""),
                "code": body.get("code", ""), "name": item_name}
        if item_id:
            m = conn.execute("SELECT * FROM price_master WHERE id=?", (item_id,)).fetchone()
            if not m:
                return err({"item_id": "選んだ商品が見つかりません"})
            snap = {"daily_rate": m["daily_rate"] or 0, "monthly_rate": m["monthly_rate"] or 0,
                    "basic_fee": m["basic_fee"] or 0, "support_per_day": m["support_per_day"] or 0,
                    "damage_per_day": m["damage_per_day"] or 0, "spec": m["spec"],
                    "code": m["code"], "name": m["name"]}
        conn.execute(
            "INSERT INTO rentals(site_id,vendor_id,item_id,item_name,spec,code,qty,"
            "start_date,due_date,daily_rate,monthly_rate,basic_fee,support_per_day,"
            "damage_per_day,created_by,client_key,created_at)"
            " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (site_id, vendor_id, item_id, snap["name"], snap["spec"], snap["code"], qty,
             start.isoformat(), due.isoformat(), snap["daily_rate"], snap["monthly_rate"],
             snap["basic_fee"], snap["support_per_day"], snap["damage_per_day"],
             user["id"], ck, now_str()))
        rid = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        for pid in body.get("photo_ids") or []:
            conn.execute("UPDATE photos SET target_type='rental', target_id=? WHERE id=? AND taken_by=?",
                         (rid, pid, user["id"]))
        touch_site_usage(conn, user["id"], site_id)
        if item_id:
            recent = get_pref(conn, user["id"], "recent_items", [])
            set_pref(conn, user["id"], "recent_items",
                     ([item_id] + [i for i in recent if i != item_id])[:10])
        audit(conn, user, "レンタル開始登録", "rental", rid,
              after={"site_id": site_id, "item": snap["name"], "qty": qty,
                     "start": start.isoformat(), "due": due.isoformat()},
              device=device_of(request))
        conn.commit()
        row = dict(conn.execute("SELECT * FROM rentals WHERE id=?", (rid,)).fetchone())
        return {"ok": True, "id": rid, "estimate": rental_snapshot_estimate(row)}
    finally:
        conn.close()


@app.get("/api/rentals")
def list_rentals(site_id: int = 0, status: str = "", mine: int = 0,
                 today: int = 0, user=Depends(current_user)):
    conn = get_db()
    sql = ("SELECT r.*, s.name site_name, v.name vendor_name FROM rentals r "
           "LEFT JOIN sites s ON s.id=r.site_id LEFT JOIN vendors v ON v.id=r.vendor_id "
           "WHERE r.deleted=0")
    args = []
    if site_id:
        sql += " AND r.site_id=?"
        args.append(site_id)
    if status:
        sql += " AND r.status=?"
        args.append(status)
    if mine or user["role"] != "admin":
        # 従業員: 自分の記録＋現場のレンタル中一覧のみ
        if not (status == "active" and site_id):
            sql += " AND r.created_by=?"
            args.append(user["id"])
    if today:
        sql += " AND r.created_at LIKE ?"
        args.append(f"{now_str()[:10]}%")
    rows = []
    for r in conn.execute(sql + " ORDER BY r.id DESC", args):
        d = dict(r)
        d["days_elapsed"] = (date.fromisoformat(now_str()[:10]) -
                             date.fromisoformat(d["start_date"])).days + 1
        d["photos"] = [dict(p) for p in conn.execute(
            "SELECT id,category,file_path FROM photos WHERE target_type='rental' "
            "AND target_id=? AND deleted=0 ORDER BY sort_order,id", (d["id"],))]
        est = rental_snapshot_estimate(d)
        d["amount_est"] = est["total"] if est else None
        rows.append(d)
    conn.close()
    return rows


@app.get("/api/rentals/{rid}/estimate")
def rental_estimate(rid: int, end: str = "", user=Depends(current_user)):
    conn = get_db()
    row = conn.execute("SELECT * FROM rentals WHERE id=? AND deleted=0", (rid,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "レンタル記録が見つかりません")
    d = dict(row)
    if end:
        d["returned_date"] = end
    return {"estimate": rental_snapshot_estimate(d)}


@app.post("/api/rentals/return")
async def return_rentals(request: Request, user=Depends(current_user)):
    body = await request.json()
    errors = {}
    ids = body.get("ids") or []
    if not ids:
        errors["ids"] = "返却する商品を選んでください"
    rdate = parse_date(body.get("returned_date"), "returned_date", errors, "返却日")
    flags = body.get("condition_flags") or {}
    if errors:
        return err(errors)
    conn = get_db()
    try:
        done = []
        for rid in ids:
            row = conn.execute("SELECT * FROM rentals WHERE id=? AND deleted=0", (rid,)).fetchone()
            if not row:
                continue
            if row["status"] == "returned":
                continue
            if rdate < date.fromisoformat(row["start_date"]):
                return err({"returned_date": "返却日は開始日以降にしてください"})
            conn.execute(
                "UPDATE rentals SET status='returned', returned_date=?, condition_flags=?,"
                "condition_comment=? WHERE id=?",
                (rdate.isoformat(), json.dumps(flags, ensure_ascii=False),
                 body.get("comment", ""), rid))
            for pid in body.get("photo_ids") or []:
                conn.execute("UPDATE photos SET target_type='rental', target_id=? "
                             "WHERE id=? AND taken_by=?", (rid, pid, user["id"]))
            audit(conn, user, "レンタル返却登録", "rental", rid,
                  before={"status": row["status"]},
                  after={"returned_date": rdate.isoformat(), "flags": flags},
                  device=device_of(request))
            done.append(rid)
        conn.commit()
        return {"ok": True, "returned": done}
    finally:
        conn.close()


@app.post("/api/rentals/{rid}/extend")
async def extend_rental(rid: int, request: Request, user=Depends(current_user)):
    body = await request.json()
    errors = {}
    new_due = parse_date(body.get("new_due"), "new_due", errors, "新しい返却予定日")
    reason = body.get("reason")
    if reason not in D.EXTENSION_REASONS:
        errors["reason"] = "延長理由を選んでください"
    if reason == "その他" and not (body.get("comment") or "").strip():
        errors["comment"] = "「その他」の場合は理由を入力してください"
    if errors:
        return err(errors)
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM rentals WHERE id=? AND deleted=0", (rid,)).fetchone()
        if not row:
            raise HTTPException(404, "レンタル記録が見つかりません")
        if row["status"] == "returned":
            return err({"new_due": "返却済みの記録は延長できません"})
        if new_due <= date.fromisoformat(row["due_date"]):
            return err({"new_due": "新しい返却予定日は現在の予定日より後にしてください"})
        conn.execute("UPDATE rentals SET due_date=? WHERE id=?", (new_due.isoformat(), rid))
        conn.execute(
            "INSERT INTO rental_extensions(rental_id,old_due,new_due,reason,comment,"
            "created_by,created_at) VALUES(?,?,?,?,?,?,?)",
            (rid, row["due_date"], new_due.isoformat(), reason,
             body.get("comment", ""), user["id"], now_str()))
        audit(conn, user, "レンタル延長登録", "rental", rid,
              before={"due_date": row["due_date"]},
              after={"due_date": new_due.isoformat(), "reason": reason},
              device=device_of(request))
        conn.commit()
        d = dict(conn.execute("SELECT * FROM rentals WHERE id=?", (rid,)).fetchone())
        d["returned_date"] = new_due.isoformat()
        return {"ok": True, "estimate": rental_snapshot_estimate(d)}
    finally:
        conn.close()


@app.put("/api/rentals/{rid}")
async def update_rental(rid: int, request: Request, user=Depends(current_user)):
    body = await request.json()
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM rentals WHERE id=? AND deleted=0", (rid,)).fetchone()
        if not row:
            raise HTTPException(404, "レンタル記録が見つかりません")
        if user["role"] != "admin" and row["created_by"] != user["id"]:
            raise HTTPException(403, "自分が登録した記録だけ修正できます")
        errors = {}
        updates = {}
        if "qty" in body:
            try:
                updates["qty"] = int(body["qty"])
                if updates["qty"] < 1:
                    errors["qty"] = "数量は1以上で入力してください"
            except (TypeError, ValueError):
                errors["qty"] = "数量は数字で入力してください"
        for f, label in [("start_date", "レンタル開始日"), ("due_date", "返却予定日"),
                         ("returned_date", "返却日")]:
            if body.get(f):
                v = parse_date(body[f], f, errors, label)
                if v:
                    updates[f] = v.isoformat()
        # 従業員は単価を変更できない
        if user["role"] == "admin":
            vals = _price_fields(body, errors)
            for f, v in vals.items():
                if body.get(f) not in (None, ""):
                    updates[f] = v
        if errors:
            return err(errors)
        if updates:
            sets = ",".join(f"{k}=?" for k in updates)
            conn.execute(f"UPDATE rentals SET {sets} WHERE id=?", (*updates.values(), rid))
            audit(conn, user, "レンタル記録修正", "rental", rid,
                  before={k: row[k] for k in updates}, after=updates,
                  device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.delete("/api/rentals/{rid}")
def delete_rental(rid: int, request: Request, user=Depends(current_user)):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM rentals WHERE id=? AND deleted=0", (rid,)).fetchone()
        if not row:
            raise HTTPException(404, "レンタル記録が見つかりません")
        if user["role"] != "admin" and row["created_by"] != user["id"]:
            raise HTTPException(403, "自分が登録した記録だけ削除できます")
        conn.execute("UPDATE rentals SET deleted=1 WHERE id=?", (rid,))
        audit(conn, user, "レンタル記録削除", "rental", rid,
              before={"item": row["item_name"], "qty": row["qty"]}, device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


# ---------------------------------------------------------------- 廃棄物
@app.post("/api/waste")
async def create_waste(request: Request, user=Depends(current_user)):
    body = await request.json()
    errors = {}
    site_id = body.get("site_id")
    if not site_id:
        errors["site_id"] = "現場を選んでください"
    out_date = parse_date(body.get("out_date"), "out_date", errors, "搬出日")
    wtype = body.get("waste_type")
    if not wtype:
        errors["waste_type"] = "廃棄物の種類を選んでください"
    try:
        qty = float(body.get("qty") or 0)
        if qty <= 0:
            errors["qty"] = "数量は0より大きい値で入力してください"
    except (TypeError, ValueError):
        errors["qty"] = "数量は数字で入力してください"
    if body.get("unit") not in D.WASTE_UNITS:
        errors["unit"] = "単位を選んでください"
    if not body.get("hauler_id") and not (body.get("hauler_name") or "").strip():
        errors["hauler_id"] = "運搬業者を選んでください"
    if not body.get("disposal_id") and not (body.get("disposal_name") or "").strip():
        errors["disposal_id"] = "処分先を選んでください"
    if errors:
        return err(errors)
    conn = get_db()
    try:
        ck = body.get("client_key")
        if ck:
            dup = conn.execute("SELECT id FROM waste_records WHERE client_key=?", (ck,)).fetchone()
            if dup:
                return {"ok": True, "id": dup["id"], "duplicate": True}

        def resolve_vendor(kind, vid_key, name_key):
            vid = body.get(vid_key)
            if vid:
                return vid
            name = body[name_key].strip()
            r = conn.execute("SELECT id FROM vendors WHERE kind=? AND name=? AND deleted=0",
                             (kind, name)).fetchone()
            if r:
                return r["id"]
            conn.execute("INSERT INTO vendors(kind,name) VALUES(?,?)", (kind, name))
            return conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]

        hauler_id = resolve_vendor("hauler", "hauler_id", "hauler_name")
        disposal_id = resolve_vendor("disposal", "disposal_id", "disposal_name")
        conn.execute(
            "INSERT INTO waste_records(site_id,out_date,waste_type,qty,unit,hauler_id,"
            "disposal_id,slip_no,created_by,client_key,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            (site_id, out_date.isoformat(), wtype, qty, body["unit"], hauler_id,
             disposal_id, body.get("slip_no", ""), user["id"], ck, now_str()))
        wid = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        for pid in body.get("photo_ids") or []:
            conn.execute("UPDATE photos SET target_type='waste', target_id=? WHERE id=? AND taken_by=?",
                         (wid, pid, user["id"]))
        touch_site_usage(conn, user["id"], site_id)
        set_pref(conn, user["id"], f"waste_defaults_{site_id}",
                 {"hauler_id": hauler_id, "disposal_id": disposal_id})
        audit(conn, user, "廃棄物搬出登録", "waste", wid,
              after={"site_id": site_id, "type": wtype, "qty": qty, "unit": body["unit"]},
              device=device_of(request))
        conn.commit()
        return {"ok": True, "id": wid}
    finally:
        conn.close()


@app.get("/api/waste")
def list_waste(site_id: int = 0, today: int = 0, user=Depends(current_user)):
    conn = get_db()
    sql = ("SELECT w.*, s.name site_name, h.name hauler_name, d2.name disposal_name "
           "FROM waste_records w LEFT JOIN sites s ON s.id=w.site_id "
           "LEFT JOIN vendors h ON h.id=w.hauler_id LEFT JOIN vendors d2 ON d2.id=w.disposal_id "
           "WHERE w.deleted=0")
    args = []
    if site_id:
        sql += " AND w.site_id=?"
        args.append(site_id)
    if user["role"] != "admin":
        sql += " AND w.created_by=?"
        args.append(user["id"])
    if today:
        sql += " AND w.created_at LIKE ?"
        args.append(f"{now_str()[:10]}%")
    rows = []
    for r in conn.execute(sql + " ORDER BY w.id DESC", args):
        d = dict(r)
        d["photos"] = [dict(p) for p in conn.execute(
            "SELECT id,category,file_path FROM photos WHERE target_type='waste' "
            "AND target_id=? AND deleted=0 ORDER BY sort_order,id", (d["id"],))]
        rows.append(d)
    conn.close()
    return rows


@app.get("/api/waste/defaults")
def waste_defaults(site_id: int, user=Depends(current_user)):
    """前回同じ現場で使用した運搬業者・処分先の候補。"""
    conn = get_db()
    try:
        d = get_pref(conn, user["id"], f"waste_defaults_{site_id}", {})
        out = {}
        for key, kind in [("hauler_id", "hauler"), ("disposal_id", "disposal")]:
            if d.get(key):
                r = conn.execute("SELECT * FROM vendors WHERE id=?", (d[key],)).fetchone()
                if r:
                    out[key] = dict(r)
        return out
    finally:
        conn.close()


@app.put("/api/waste/{wid}")
async def update_waste(wid: int, request: Request, user=Depends(current_user)):
    body = await request.json()
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM waste_records WHERE id=? AND deleted=0", (wid,)).fetchone()
        if not row:
            raise HTTPException(404, "廃棄物記録が見つかりません")
        if user["role"] != "admin" and row["created_by"] != user["id"]:
            raise HTTPException(403, "自分が登録した記録だけ修正できます")
        errors = {}
        updates = {}
        if body.get("out_date"):
            v = parse_date(body["out_date"], "out_date", errors, "搬出日")
            if v:
                updates["out_date"] = v.isoformat()
        if "qty" in body:
            try:
                updates["qty"] = float(body["qty"])
                if updates["qty"] <= 0:
                    errors["qty"] = "数量は0より大きい値で入力してください"
            except (TypeError, ValueError):
                errors["qty"] = "数量は数字で入力してください"
        for f in ("waste_type", "unit", "slip_no"):
            if body.get(f) is not None:
                updates[f] = body[f]
        if user["role"] == "admin" and body.get("amount") not in (None, ""):
            try:
                updates["amount"] = int(body["amount"])
            except (TypeError, ValueError):
                errors["amount"] = "金額は数字で入力してください"
        if errors:
            return err(errors)
        if updates:
            sets = ",".join(f"{k}=?" for k in updates)
            conn.execute(f"UPDATE waste_records SET {sets} WHERE id=?", (*updates.values(), wid))
            audit(conn, user, "廃棄物記録修正", "waste", wid,
                  before={k: row[k] for k in updates}, after=updates,
                  device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.delete("/api/waste/{wid}")
def delete_waste(wid: int, request: Request, user=Depends(current_user)):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM waste_records WHERE id=? AND deleted=0", (wid,)).fetchone()
        if not row:
            raise HTTPException(404, "廃棄物記録が見つかりません")
        if user["role"] != "admin" and row["created_by"] != user["id"]:
            raise HTTPException(403, "自分が登録した記録だけ削除できます")
        conn.execute("UPDATE waste_records SET deleted=1 WHERE id=?", (wid,))
        audit(conn, user, "廃棄物記録削除", "waste", wid,
              before={"type": row["waste_type"], "qty": row["qty"]}, device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


# ---------------------------------------------------------------- 写真
ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic", ".pdf"}


@app.post("/api/photos")
async def upload_photo(request: Request, file: UploadFile = File(...),
                       target_type: str = Form("pending"), target_id: int = Form(0),
                       category: str = Form("その他"), user=Depends(current_user)):
    ext = os.path.splitext(file.filename or "")[1].lower() or ".jpg"
    if ext not in ALLOWED_EXT:
        return err({"file": "対応していないファイル形式です（写真またはPDFを選んでください）"})
    fname = f"{now_str()[:10]}_{secrets.token_hex(8)}{ext}"
    path = os.path.join(UPLOADS, fname)
    with open(path, "wb") as f:
        f.write(await file.read())
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO photos(target_type,target_id,category,file_path,taken_by,created_at)"
            " VALUES(?,?,?,?,?,?)",
            (target_type, target_id or None, category, f"/uploads/{fname}", user["id"], now_str()))
        pid = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        audit(conn, user, "写真登録", target_type, target_id or pid,
              after={"category": category}, device=device_of(request))
        conn.commit()
        return {"ok": True, "id": pid, "url": f"/uploads/{fname}"}
    finally:
        conn.close()


@app.get("/api/photos")
def list_photos(target_type: str = "", target_id: int = 0, user=Depends(current_user)):
    conn = get_db()
    sql = "SELECT * FROM photos WHERE deleted=0"
    args = []
    if target_type:
        sql += " AND target_type=?"
        args.append(target_type)
    if target_id:
        sql += " AND target_id=?"
        args.append(target_id)
    if user["role"] != "admin" and not target_id:
        sql += " AND taken_by=?"
        args.append(user["id"])
    rows = [dict(r) for r in conn.execute(sql + " ORDER BY sort_order,id", args)]
    conn.close()
    return rows


@app.delete("/api/photos/{pid}")
def delete_photo(pid: int, request: Request, user=Depends(current_user)):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM photos WHERE id=?", (pid,)).fetchone()
        if not row:
            raise HTTPException(404, "写真が見つかりません")
        if user["role"] != "admin" and row["taken_by"] != user["id"]:
            raise HTTPException(403, "自分が撮影した写真だけ削除できます")
        conn.execute("UPDATE photos SET deleted=1 WHERE id=?", (pid,))
        audit(conn, user, "写真削除", row["target_type"], row["target_id"],
              device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


@app.post("/api/photos/reorder")
async def reorder_photos(request: Request, user=Depends(current_user)):
    body = await request.json()
    conn = get_db()
    for i, pid in enumerate(body.get("ids") or []):
        conn.execute("UPDATE photos SET sort_order=? WHERE id=?", (i, pid))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---------------------------------------------------------------- 下書き
@app.get("/api/drafts")
def list_drafts(user=Depends(current_user)):
    conn = get_db()
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM drafts WHERE user_id=? ORDER BY updated_at DESC", (user["id"],))]
    conn.close()
    return rows


@app.post("/api/drafts")
async def save_draft(request: Request, user=Depends(current_user)):
    body = await request.json()
    conn = get_db()
    try:
        did = body.get("id")
        payload = json.dumps(body.get("payload") or {}, ensure_ascii=False)
        if did:
            conn.execute("UPDATE drafts SET payload=?, updated_at=? WHERE id=? AND user_id=?",
                         (payload, now_str(), did, user["id"]))
        else:
            conn.execute("INSERT INTO drafts(user_id,kind,payload,updated_at) VALUES(?,?,?,?)",
                         (user["id"], body.get("kind", "rental"), payload, now_str()))
            did = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        conn.commit()
        return {"ok": True, "id": did}
    finally:
        conn.close()


@app.delete("/api/drafts/{did}")
def delete_draft(did: int, user=Depends(current_user)):
    conn = get_db()
    conn.execute("DELETE FROM drafts WHERE id=? AND user_id=?", (did, user["id"]))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---------------------------------------------------------------- ホーム
@app.get("/api/home")
def home(user=Depends(current_user)):
    conn = get_db()
    try:
        today = now_str()[:10]
        uid = user["id"]
        rentals_today = conn.execute(
            "SELECT COUNT(*) c FROM rentals WHERE created_by=? AND created_at LIKE ? AND deleted=0",
            (uid, f"{today}%")).fetchone()["c"]
        waste_today = conn.execute(
            "SELECT COUNT(*) c FROM waste_records WHERE created_by=? AND created_at LIKE ? AND deleted=0",
            (uid, f"{today}%")).fetchone()["c"]
        due_soon = conn.execute(
            "SELECT COUNT(*) c FROM rentals WHERE status='active' AND deleted=0 "
            "AND due_date <= date(?, '+3 day')", (today,)).fetchone()["c"]
        drafts = conn.execute("SELECT COUNT(*) c FROM drafts WHERE user_id=?", (uid,)).fetchone()["c"]
        last_site_id = get_pref(conn, uid, "last_site")
        last_site = None
        if last_site_id:
            r = conn.execute("SELECT * FROM sites WHERE id=? AND deleted=0", (last_site_id,)).fetchone()
            last_site = dict(r) if r else None
        return {"user": {"display_name": user["display_name"], "role": user["role"]},
                "now": now_str(), "last_site": last_site,
                "today_count": rentals_today + waste_today, "due_soon": due_soon,
                "drafts": drafts}
    finally:
        conn.close()


# ---------------------------------------------------------------- 管理者: 集計・出力
def month_rows(conn, month: str):
    """month=YYYY-MM のレンタル月別内訳と廃棄物。"""
    y, m = int(month[:4]), int(month[5:7])
    rentals = []
    for r in conn.execute(
            "SELECT r.*, s.name site_name, v.name vendor_name FROM rentals r "
            "LEFT JOIN sites s ON s.id=r.site_id LEFT JOIN vendors v ON v.id=r.vendor_id "
            "WHERE r.deleted=0"):
        est = rental_snapshot_estimate(dict(r))
        if not est:
            continue
        for line in est["monthly"]:
            if line["year"] == y and line["month"] == m:
                d = dict(r)
                d["month_line"] = line
                rentals.append(d)
    waste = [dict(r) for r in conn.execute(
        "SELECT w.*, s.name site_name, h.name hauler_name, d2.name disposal_name "
        "FROM waste_records w LEFT JOIN sites s ON s.id=w.site_id "
        "LEFT JOIN vendors h ON h.id=w.hauler_id LEFT JOIN vendors d2 ON d2.id=w.disposal_id "
        "WHERE w.deleted=0 AND w.out_date LIKE ?", (f"{month}%",))]
    return rentals, waste


@app.get("/api/admin/summary")
def admin_summary(month: str, user=Depends(admin_user)):
    if not re.match(r"^\d{4}-\d{2}$", month):
        raise HTTPException(400, "月は YYYY-MM 形式で指定してください")
    conn = get_db()
    try:
        rentals, waste = month_rows(conn, month)
        sites: dict = {}
        for r in rentals:
            s = sites.setdefault(r["site_name"] or "(現場未設定)",
                                 {"rental_total": 0, "waste_total": 0, "rentals": [], "waste": []})
            s["rental_total"] += r["month_line"]["subtotal"]
            s["rentals"].append({
                "id": r["id"], "item_name": r["item_name"], "qty": r["qty"],
                "start_date": r["start_date"], "due_date": r["due_date"],
                "returned_date": r["returned_date"],
                **r["month_line"]})
        for w in waste:
            s = sites.setdefault(w["site_name"] or "(現場未設定)",
                                 {"rental_total": 0, "waste_total": 0, "rentals": [], "waste": []})
            s["waste_total"] += w["amount"] or 0
            s["waste"].append({k: w[k] for k in ("id", "out_date", "waste_type", "qty", "unit",
                                                 "hauler_name", "disposal_name", "amount")})
        total = sum(s["rental_total"] + s["waste_total"] for s in sites.values())
        return {"month": month, "sites": sites, "grand_total": total}
    finally:
        conn.close()


@app.get("/api/admin/export/csv")
def export_csv(month: str, user=Depends(admin_user)):
    conn = get_db()
    rentals, waste = month_rows(conn, month)
    conn.close()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["種別", "現場", "品名/廃棄物", "規格", "数量", "単位", "開始日", "返却予定日",
                "返却日", "当月日数", "レンタル料", "基本料", "サポート料", "賠償対策費",
                "当月小計"])
    for r in rentals:
        ml = r["month_line"]
        w.writerow(["レンタル", r["site_name"], r["item_name"], r["spec"], r["qty"], "台",
                    r["start_date"], r["due_date"], r["returned_date"] or "", ml["days"],
                    ml["rental"], ml["basic"], ml["support"], ml["damage"], ml["subtotal"]])
    for x in waste:
        w.writerow(["廃棄物", x["site_name"], x["waste_type"], "", x["qty"], x["unit"],
                    x["out_date"], "", "", "", "", "", "", "", x["amount"] or ""])
    data = "﻿" + buf.getvalue()  # Excelで文字化けしないようBOM付き
    return Response(data, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=summary_{month}.csv"})


@app.get("/api/admin/export/xlsx")
def export_xlsx(month: str, user=Depends(admin_user)):
    from openpyxl import Workbook
    conn = get_db()
    rentals, waste = month_rows(conn, month)
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month} レンタル"
    ws.append(["現場", "品名", "規格", "数量", "開始日", "返却予定日", "返却日", "当月日数",
               "レンタル料", "基本料", "サポート料", "賠償対策費", "当月小計"])
    for r in rentals:
        ml = r["month_line"]
        ws.append([r["site_name"], r["item_name"], r["spec"], r["qty"], r["start_date"],
                   r["due_date"], r["returned_date"] or "", ml["days"], ml["rental"],
                   ml["basic"], ml["support"], ml["damage"], ml["subtotal"]])
    ws2 = wb.create_sheet(f"{month} 廃棄物")
    ws2.append(["現場", "搬出日", "種類", "数量", "単位", "運搬業者", "処分先", "金額"])
    for x in waste:
        ws2.append([x["site_name"], x["out_date"], x["waste_type"], x["qty"], x["unit"],
                    x["hauler_name"], x["disposal_name"], x["amount"] or ""])
    out = io.BytesIO()
    wb.save(out)
    return Response(out.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=summary_{month}.xlsx"})


@app.get("/api/admin/export/pdf")
def export_pdf(month: str, user=Depends(admin_user)):
    """印刷用HTML（ブラウザの印刷→PDF保存で利用）。"""
    conn = get_db()
    rentals, waste = month_rows(conn, month)
    company = conn.execute("SELECT * FROM company WHERE id=1").fetchone()
    conn.close()
    rows = "".join(
        f"<tr><td>{r['site_name']}</td><td>{r['item_name']} {r['spec']}</td><td>{r['qty']}</td>"
        f"<td>{r['start_date']}〜{r['returned_date'] or r['due_date']}</td>"
        f"<td style='text-align:right'>{r['month_line']['subtotal']:,}円</td></tr>"
        for r in rentals)
    wrows = "".join(
        f"<tr><td>{x['site_name']}</td><td>{x['waste_type']}</td><td>{x['qty']}{x['unit']}</td>"
        f"<td>{x['out_date']}</td><td style='text-align:right'>{(x['amount'] or 0):,}円</td></tr>"
        for x in waste)
    total = sum(r["month_line"]["subtotal"] for r in rentals) + sum(x["amount"] or 0 for x in waste)
    html = f"""<!doctype html><html lang="ja"><meta charset="utf-8">
<title>{month} 月次集計</title>
<style>body{{font-family:sans-serif;margin:24px}}table{{border-collapse:collapse;width:100%;margin:12px 0}}
td,th{{border:1px solid #999;padding:6px;font-size:13px}}h1{{font-size:20px}}
@media print{{button{{display:none}}}}</style>
<button onclick="print()">印刷 / PDF保存</button>
<h1>{(company['company_name'] if company else '')} 月次集計 {month}</h1>
<h2>レンタル</h2><table><tr><th>現場</th><th>商品</th><th>数量</th><th>期間</th><th>当月小計</th></tr>{rows or '<tr><td colspan=5>記録なし</td></tr>'}</table>
<h2>廃棄物</h2><table><tr><th>現場</th><th>種類</th><th>数量</th><th>搬出日</th><th>金額</th></tr>{wrows or '<tr><td colspan=5>記録なし</td></tr>'}</table>
<h2 style='text-align:right'>合計 {total:,}円</h2></html>"""
    return HTMLResponse(html)


@app.get("/api/admin/backup")
def backup(user=Depends(admin_user)):
    conn = get_db()
    tables = ["users", "company", "sites", "vendors", "price_master", "rentals",
              "rental_extensions", "waste_records", "photos", "delete_requests",
              "audit_log", "user_prefs", "price_import_batches", "price_import_rows"]
    data = {}
    for t in tables:
        rows = [dict(r) for r in conn.execute(f"SELECT * FROM {t}")]
        if t == "users":
            for r in rows:
                r.pop("password_hash", None)
        data[t] = rows
    conn.close()
    return Response(json.dumps(data, ensure_ascii=False, indent=1),
                    media_type="application/json",
                    headers={"Content-Disposition":
                             f"attachment; filename=backup_{now_str()[:10]}.json"})


@app.get("/api/admin/audit")
def audit_list(limit: int = 200, user=Depends(admin_user)):
    conn = get_db()
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM audit_log ORDER BY id DESC LIMIT ?", (min(limit, 1000),))]
    conn.close()
    return rows


@app.get("/api/admin/company")
def get_company(user=Depends(admin_user)):
    conn = get_db()
    row = conn.execute("SELECT * FROM company WHERE id=1").fetchone()
    conn.close()
    return dict(row) if row else {}


@app.put("/api/admin/company")
async def update_company(request: Request, user=Depends(admin_user)):
    body = await request.json()
    conn = get_db()
    try:
        before = conn.execute("SELECT * FROM company WHERE id=1").fetchone()
        conn.execute("UPDATE company SET company_name=?, closing_day=? WHERE id=1",
                     (body.get("company_name", before["company_name"] if before else ""),
                      body.get("closing_day", before["closing_day"] if before else 31)))
        audit(conn, user, "会社設定変更", "company", 1,
              before=dict(before) if before else None, after=body, device=device_of(request))
        conn.commit()
        return {"ok": True}
    finally:
        conn.close()


# ---------------------------------------------------------------- 単価表取込
def _to_int_or_none(v):
    if v in (None, ""):
        return None
    try:
        return int(round(float(str(v).replace(",", "").replace("¥", "").replace("円", ""))))
    except (TypeError, ValueError):
        return None


IMPORT_COLS = {"品名": "name", "商品コード": "code", "規格": "spec",
               "日割単価": "daily_rate", "月極単価": "monthly_rate",
               "サポート料": "support_per_day", "サポート料／日": "support_per_day",
               "基本料": "basic_fee", "賠償対策費": "damage_per_day",
               "賠償対策費／日": "damage_per_day"}


@app.post("/api/admin/import/upload")
async def import_upload(request: Request, file: UploadFile = File(...),
                        user=Depends(admin_user)):
    ext = os.path.splitext(file.filename or "")[1].lower()
    fname = f"import_{secrets.token_hex(6)}{ext}"
    path = os.path.join(UPLOADS, fname)
    content = await file.read()
    with open(path, "wb") as f:
        f.write(content)
    conn = get_db()
    try:
        conn.execute("INSERT INTO price_import_batches(filename,image_path,created_by,created_at)"
                     " VALUES(?,?,?,?)", (file.filename, f"/uploads/{fname}", user["id"], now_str()))
        bid = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        note = ""
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in ws[1]]
            mapping = {i: IMPORT_COLS[h] for i, h in enumerate(headers) if h in IMPORT_COLS}
            if "name" not in mapping.values():
                note = "1行目に「品名」列が見つかりません。列名を確認してください"
            else:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rec = {"name": "", "code": "", "spec": ""}
                    nums = {}
                    for i, v in enumerate(row):
                        f2 = mapping.get(i)
                        if not f2:
                            continue
                        if f2 in ("name", "code", "spec"):
                            rec[f2] = str(v or "").strip()
                        else:
                            nums[f2] = _to_int_or_none(v)
                            if v not in (None, "") and nums[f2] is None:
                                nums[f2 + "_raw_bad"] = True
                    if not rec["name"]:
                        continue
                    # 空欄・読み取れない数値は補完せず「要確認」
                    needs = any(nums.get(k) is None for k in
                                ("daily_rate", "monthly_rate", "basic_fee",
                                 "support_per_day", "damage_per_day"))
                    conn.execute(
                        "INSERT INTO price_import_rows(batch_id,code,name,spec,daily_rate,"
                        "monthly_rate,basic_fee,support_per_day,damage_per_day,needs_review,note)"
                        " VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                        (bid, rec["code"], rec["name"], rec["spec"],
                         nums.get("daily_rate"), nums.get("monthly_rate"),
                         nums.get("basic_fee"), nums.get("support_per_day"),
                         nums.get("damage_per_day"), 1 if needs else 0,
                         "空欄または読み取れない数値があります（要確認）" if needs else ""))
        else:
            note = ("画像を取り込みました。自動読取が利用できない環境のため、"
                    "原本画像を見ながら行を追加して確認してください。"
                    "読み取れない数値は空欄のまま「要確認」にしてください")
        audit(conn, user, "単価表取込アップロード", "price_import_batch", bid,
              after={"filename": file.filename}, device=device_of(request))
        conn.commit()
        return {"ok": True, "batch_id": bid, "note": note}
    finally:
        conn.close()


@app.get("/api/admin/import/batches")
def import_batches(user=Depends(admin_user)):
    conn = get_db()
    rows = []
    for b in conn.execute("SELECT * FROM price_import_batches ORDER BY id DESC"):
        d = dict(b)
        d["pending"] = conn.execute(
            "SELECT COUNT(*) c FROM price_import_rows WHERE batch_id=? AND status='pending'",
            (b["id"],)).fetchone()["c"]
        rows.append(d)
    conn.close()
    return rows


@app.get("/api/admin/import/batches/{bid}")
def import_batch(bid: int, user=Depends(admin_user)):
    conn = get_db()
    b = conn.execute("SELECT * FROM price_import_batches WHERE id=?", (bid,)).fetchone()
    if not b:
        conn.close()
        raise HTTPException(404, "取込データが見つかりません")
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM price_import_rows WHERE batch_id=? ORDER BY id", (bid,))]
    conn.close()
    return {"batch": dict(b), "rows": rows}


@app.post("/api/admin/import/rows")
async def import_row_add(request: Request, user=Depends(admin_user)):
    body = await request.json()
    if not body.get("batch_id"):
        return err({"batch_id": "取込データを選んでください"})
    if not (body.get("name") or "").strip():
        return err({"name": "品名を入力してください"})
    errors = {}
    vals = _price_fields(body, errors)
    if errors:
        return err(errors)
    needs = 1 if any(vals[f] is None for f in vals) else 0
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO price_import_rows(batch_id,code,name,spec,daily_rate,monthly_rate,"
            "basic_fee,support_per_day,damage_per_day,needs_review,note) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
            (body["batch_id"], body.get("code", ""), body["name"].strip(), body.get("spec", ""),
             vals["daily_rate"], vals["monthly_rate"], vals["basic_fee"],
             vals["support_per_day"], vals["damage_per_day"], needs,
             "空欄があります（要確認）" if needs else ""))
        rid = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        conn.commit()
        return {"ok": True, "id": rid, "needs_review": needs}
    finally:
        conn.close()


@app.put("/api/admin/import/rows/{rid}")
async def import_row_edit(rid: int, request: Request, user=Depends(admin_user)):
    body = await request.json()
    errors = {}
    vals = _price_fields(body, errors)
    if errors:
        return err(errors)
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM price_import_rows WHERE id=?", (rid,)).fetchone()
        if not row:
            raise HTTPException(404, "行が見つかりません")
        needs = 1 if any(vals[f] is None for f in vals) else 0
        conn.execute(
            "UPDATE price_import_rows SET code=?,name=?,spec=?,daily_rate=?,monthly_rate=?,"
            "basic_fee=?,support_per_day=?,damage_per_day=?,needs_review=?,note=? WHERE id=?",
            (body.get("code", row["code"]), body.get("name", row["name"]),
             body.get("spec", row["spec"]), vals["daily_rate"], vals["monthly_rate"],
             vals["basic_fee"], vals["support_per_day"], vals["damage_per_day"],
             needs, "" if not needs else "空欄があります（要確認）", rid))
        conn.commit()
        return {"ok": True, "needs_review": needs}
    finally:
        conn.close()


@app.post("/api/admin/import/rows/{rid}/approve")
def import_row_approve(rid: int, request: Request, user=Depends(admin_user)):
    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM price_import_rows WHERE id=?", (rid,)).fetchone()
        if not row:
            raise HTTPException(404, "行が見つかりません")
        if row["needs_review"]:
            return err({"row": "要確認の項目が残っています。数値を確認してから承認してください"})
        batch = conn.execute("SELECT * FROM price_import_batches WHERE id=?",
                             (row["batch_id"],)).fetchone()
        conn.execute(
            "INSERT INTO price_master(code,name,spec,daily_rate,monthly_rate,basic_fee,"
            "support_per_day,damage_per_day,needs_review,source_image,created_at)"
            " VALUES(?,?,?,?,?,?,?,?,0,?,?)",
            (row["code"], row["name"], row["spec"], row["daily_rate"], row["monthly_rate"],
             row["basic_fee"], row["support_per_day"], row["damage_per_day"],
             batch["image_path"] if batch else "", now_str()))
        pid = conn.execute("SELECT last_insert_rowid() i").fetchone()["i"]
        conn.execute("UPDATE price_import_rows SET status='approved' WHERE id=?", (rid,))
        audit(conn, user, "単価表取込承認", "price_master", pid,
              after={"name": row["name"], "code": row["code"]}, device=device_of(request))
        conn.commit()
        return {"ok": True, "price_master_id": pid}
    finally:
        conn.close()


@app.post("/api/admin/import/rows/{rid}/reject")
def import_row_reject(rid: int, user=Depends(admin_user)):
    conn = get_db()
    conn.execute("UPDATE price_import_rows SET status='rejected' WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---------------------------------------------------------------- 静的ファイル
app.mount("/uploads", StaticFiles(directory=UPLOADS), name="uploads")


@app.get("/", response_class=HTMLResponse)
def index():
    return FileResponse(os.path.join(STATIC, "index.html"))


app.mount("/static", StaticFiles(directory=STATIC), name="static")
